from ctypes import pythonapi
from tika import parser
import os, sys
from PyPDF2 import PdfFileWriter, PdfFileReader
from openpyxl import Workbook 
from tkinter import Tk
from tkinter.filedialog import askopenfilename

Tk().withdraw()
filename = askopenfilename()
print(filename)

PDF_file = filename

inputpdf = PdfFileReader(open(PDF_file, "rb"))

wb = Workbook()
ws =  wb.active
ws.title = "Notas Fiscais"
ws['A1'] = 'DATA EMISSÃO'
ws['B1'] = 'Nº NF'
ws['C1'] = 'SERVIÇO'
ws['D1'] = 'CNPJ CLIENTE'
ws['E1'] = 'NOME'
ws['F1'] = 'VALOR BRUTO'
ws['G1'] = 'IR'
ws['H1'] = 'INSS'
ws['I1'] = 'CSLL'
ws['J1'] = 'COFINS'
ws['K1'] = 'PIS'

wb.save(filename = f'{os.path.dirname(sys.argv[0])}/Notas Fiscais.xlsx')

def search( dict, first, last ):
    str = list(dict.values())[1]
    start = str.find( first ) + len( first )
    end = str.find( last, start )
    index = 1
    while(end == -1 and index < len(last) ):
        end = str.find( last[index], start )
        index= index + 1
    if end !=-1:
        return str[start:end]
    else:
        return ''
        
for i in range(inputpdf.numPages):
    output = PdfFileWriter()
    output.addPage(inputpdf.getPage(i))
    with open(f'{os.path.dirname(sys.argv[0])}/document-page%s.pdf' % i, "wb") as outputStream:
        output.write(outputStream)
    
    raw = parser.from_file(f'{os.path.dirname(sys.argv[0])}/document-page%s.pdf' % i) 

    ws =  wb.worksheets[0]

    first = 'Valor Total'
    last = 'Valor Líquido'
    valor = search(raw, first, last).replace('\n','')
    x=i+2
    ws.cell(column = 6,row = i + 2 ).value = valor

    first = 'Data Emissão'
    last = 'Hora Emissão'
    data = search(raw, first, last).replace('\n','')
    ws['A%s' %x] = data

    first = 'Número da NFS-e'
    last = 'Situação'
    numero = search(raw, first, last).replace('\n','')
    ws['B%s' %x] = numero

    first = 'Descrição do Serviço:'
    last = 'Base de Cálculo'
    servico = search(raw, first, last).replace('\n','')
    ws['C%s' %x] = servico

    first = 'CPF/CNPJ'
    last = 'Endereço'
    cnpj = search(raw, first, last).replace('\n','')
    ws['D%s' %x] = cnpj

    first = 'TOMADOR DO SERVIÇO\nRazão Social'
    last = 'CPF/CNPJ'
    nome = search(raw, first, last).replace('\n','')
    ws['E%s' %x] = nome

    first = '\n\nIR\n'
    last = '\n\nINSS\n'
    ir = search(raw, first, last).replace('\n','')
    ws['G%s' %x] = ir

    first = '\n\nINSS\n'
    last = 'n\nCSLL\n'
    inss = search(raw, first, last).replace('\n','')
    ws['H%s' %x] = inss

    first = 'CSLL'
    last = 'COFINS'
    csll = search(raw, first, last).replace('\n','')
    ws['I%s' %x] = csll

    first = 'COFINS'
    last = '\n\nPIS\n'
    cofins = search(raw, first, last).replace('\n','')
    ws['J%s' %x] = cofins

    first = '\n\nPIS\n'
    last = 'Descrição dos subitens'
    pis = search(raw, first, last).replace('\n','')
    ws['K%s' %x] = pis   

    #print(nome,'\n', cnpj,'\n', servico,'\n', numero,'\n', data,'\n', valor)

    os.remove(f'{os.path.dirname(sys.argv[0])}/document-page%s.pdf' % i)

    print('%s de %s PDFs' %(i+1, len(range(inputpdf.numPages))))
wb.save(filename = f'{os.path.dirname(sys.argv[0])}/Notas Fiscais.xlsx')

print('Importação concluída')   

#print(list(raw.values()))
